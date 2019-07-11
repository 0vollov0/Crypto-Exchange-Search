from bs4 import BeautifulSoup
from pprint import pprint
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
import sys

class CoinMarketResearch:
    def __init__(self):
        self.markets = []
        self.load_wb_market_names = ''
        self.load_wb_previous_researcher = ''
        self.previous_researcher_list = []
        self.load_files()
        self.register_market_name()
        self.register_previous_researcher_list()

    def load_files(self):
        #거래소 파일 불러오기
        try:
            print('거래소 이름 파일의 절대경로를 입력해주세요.')
            self.load_wb_market_names = load_workbook(str(input()), data_only=True)
            print('이전 거래소 조사 파일의 절대경로를 입력해주세요.')
            self.load_wb_previous_researcher = load_workbook(str(input()), data_only=True)

            self.load_ws_market_names = self.load_wb_market_names['total']
            self.load_wb_previous_researcher = self.load_wb_previous_researcher['total']
        except FileNotFoundError:
            print('파일을 찾을 수 없습니다.')
            sys.exit()

    def register_market_name(self):
        for index,row in enumerate(self.load_ws_market_names):
            self.markets.append(self.load_ws_market_names.cell(index+1,1).value)

    def register_previous_researcher_list(self):
        #이전 거래소 조사 정보
        index = 0;
        for data in self.load_wb_previous_researcher:
            index+=1

        get_cells = self.load_wb_previous_researcher['C2':'E'+str(index)]

        for row in get_cells:
            previous_researcher = {}
            previous_researcher['Bitcoin'] = row[0].value
            previous_researcher['Ethereum'] = row[1].value
            previous_researcher['XRP'] = row[2].value
            self.previous_researcher_list.append(previous_researcher)


    def research_coinMarket(self):
        result_list = []
        for market in self.markets:
            html = requests.get('https://coinmarketcap.com/ko/exchanges/'+market)
            soup = BeautifulSoup(html.text,'html.parser')
            coin_dic = {}

            market_table = soup.find('table',{'id' : 'exchange-markets'})
            if  market_table == None:
                self.closed_market(result_list, market)
                continue

            table_rows = market_table.find('tbody').findAll('tr')


            row_list = []
            for row in table_rows:
                row_list.append(row)

            coin_list = []

            for coin in row_list:
                coin_list.append(coin.find('a',{'class':'margin-left--lv1 link-secondary'}).text)
            coin_dic['market'] = market

            if 'Bitcoin' in coin_list:
                coin_dic['Bitcoin'] = 1
            else:
                coin_dic['Bitcoin'] = 0

            if 'Ethereum' in coin_list:
                coin_dic['Ethereum'] = 1
            else:
                coin_dic['Ethereum'] = 0

            if 'XRP' in coin_list:
                coin_dic['XRP'] = 1
            else:
                coin_dic['XRP'] = 0
            result_list.append(coin_dic)
        return result_list

    def closed_market(self,result_list, market):
        result_list.append({'market': market, 'Bitcoin': 0, 'Ethereum': 0, 'XRP': 0})

    def init_sheet(self,sheet):
        sheet['A1'] = 'idx'
        sheet['B1'] = 'exhange_name'
        sheet['C1'] = 'btc'
        sheet['D1'] = 'eth'
        sheet['E1'] = 'xrp'
        sheet['F1'] = 'explain'

    def write_market_research(self):
        coin_dic = self.research_coinMarket()

        write_wb = Workbook()

        write_ws_sheet_total = write_wb.create_sheet("total",0)
        write_ws_sheet_not_changed = write_wb.create_sheet("Not_Changed",1)
        write_ws_sheet_changed = write_wb.create_sheet("Changed",2)
        write_ws_sheet_closed = write_wb.create_sheet("Closed",3)

        self.init_sheet(write_ws_sheet_total)
        self.init_sheet(write_ws_sheet_not_changed)
        self.init_sheet(write_ws_sheet_changed)
        self.init_sheet(write_ws_sheet_closed)

        for index,coin in enumerate(coin_dic):
            bit_changed  = not(coin['Bitcoin'] == self.previous_researcher_list[index]['Bitcoin'])
            eth_changed  = not(coin['Ethereum'] == self.previous_researcher_list[index]['Ethereum'])
            xrp_changed  = not(coin['XRP'] == self.previous_researcher_list[index]['XRP'])
            changed_result = ''
            if bit_changed:
                changed_result += 'Bitcoin_Changed '+str(self.previous_researcher_list[index]['Bitcoin'])+'->'+str(coin['Bitcoin'])
            if eth_changed:
                changed_result += 'Ethereum_Changed '+str(self.previous_researcher_list[index]['Ethereum'])+'->'+str(coin['Ethereum'])
            if xrp_changed:
                changed_result += 'XRP_Changed'+str(self.previous_researcher_list[index]['XRP'])+'->'+str(coin['XRP'])
            if not bit_changed and not eth_changed and not xrp_changed:
                changed_result = 'Not_Changed'
            if coin['Bitcoin'] == 0 and coin['Ethereum'] == 0 and coin['XRP'] == 0:
                changed_result += ' 거래소 폐쇄'
            if changed_result == 'Not_Changed':
                write_ws_sheet_not_changed.append([index+1,coin['market'],coin['Bitcoin'],coin['Ethereum'],coin['XRP']])
            else:
                write_ws_sheet_changed.append([index+1,coin['market'],coin['Bitcoin'],coin['Ethereum'],coin['XRP'],changed_result])
                if '거래소 폐쇄' in changed_result:
                    write_ws_sheet_closed.append([index+1,coin['market'],coin['Bitcoin'],coin['Ethereum'],coin['XRP'],changed_result])
            write_ws_sheet_total.append([index+1,coin['market'],coin['Bitcoin'],coin['Ethereum'],coin['XRP'],changed_result])

        write_wb.save('C:/Users/user/Desktop/인턴/거래소 리스트 조사_test '+time.strftime('%Y_%m_%d', time.localtime(time.time()))+'.xlsx')
        write_wb.close()
